import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import time
import random
import logging
import ssl
import certifi
import os
from typing import List, Tuple, Set, Dict, Optional
import json
import argparse
import urllib.parse
import re
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Fix SSL certificate verification
ssl._create_default_https_context = ssl._create_unverified_context

USER_AGENTS = [
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:123.0) Gecko/20100101 Firefox/123.0'
]

def format_phone_number(phone: str) -> str:
    """Format phone number according to specific rules."""
    if not phone or phone == "N/A":
        return "N/A"
    
    # Remove all spaces and special characters
    cleaned = re.sub(r'[^\d]', '', phone)
    
    # If number starts with 06, 07, 08, or 09 (mobile numbers)
    if re.match(r'^0[6-9]', cleaned):
        # Remove leading 0 and return the rest
        return cleaned[1:]
    
    # If number is in format like 02066020121 (11 digits starting with 0)
    if len(cleaned) == 11 and cleaned.startswith('0'):
        # Keep leading 0 and add hyphen after area code (first 3 digits)
        return f"{cleaned[:3]}-{cleaned[3:]}"
    
    return phone  # Return original if no rules match

def convert_review_count(review_text: str) -> str:
    """Convert review count with K to numeric format."""
    if not review_text or review_text == "N/A":
        return "N/A"
    
    # Remove any non-numeric characters except K, commas, and decimal points
    cleaned = ''.join(c for c in review_text if c.isdigit() or c == ',' or c == '.' or c.lower() == 'k')
    
    # If the text contains 'K'
    if 'k' in cleaned.lower():
        # Extract the number before K
        number = cleaned.lower().replace('k', '').replace(',', '')
        try:
            # Convert to float and multiply by 1000
            numeric_value = float(number) * 1000
            # Convert to integer and format with commas
            return f"{int(numeric_value):,}"
        except ValueError:
            return review_text
    
    return cleaned

class MouseMovement:
    def __init__(self, driver):
        self.driver = driver
        self.last_x = 0
        self.last_y = 0
        self.action = ActionChains(driver)

    def bezier_curve(self, start: Tuple[int, int], end: Tuple[int, int], control_points: int = 2) -> List[Tuple[int, int]]:
        """Generate points along a Bezier curve for smooth mouse movement."""
        points = []
        for t in range(0, 101, 5):  # Generate points at 5% intervals
            t = t / 100
            x = y = 0
            for i in range(control_points + 1):
                # Calculate binomial coefficient
                binom = math.comb(control_points, i)
                # Calculate point on curve
                x += binom * (1 - t) ** (control_points - i) * t ** i * (start[0] if i == 0 else end[0] if i == control_points else random.randint(min(start[0], end[0]), max(start[0], end[0])))
                y += binom * (1 - t) ** (control_points - i) * t ** i * (start[1] if i == 0 else end[1] if i == control_points else random.randint(min(start[1], end[1]), max(start[1], end[1])))
            points.append((int(x), int(y)))
        return points

    def move_to_element(self, element, duration: float = 1.0):
        """Move mouse to element using Bezier curve."""
        try:
            # Get element location
            location = element.location
            size = element.size
            
            # Calculate target point (center of element with some randomness)
            target_x = location['x'] + size['width'] // 2 + random.randint(-10, 10)
            target_y = location['y'] + size['height'] // 2 + random.randint(-10, 10)
            
            # Get current mouse position
            current_x = self.last_x if self.last_x != 0 else random.randint(0, 100)
            current_y = self.last_y if self.last_y != 0 else random.randint(0, 100)
            
            # Generate points along Bezier curve
            points = self.bezier_curve((current_x, current_y), (target_x, target_y))
            
            # Move through points
            for x, y in points:
                self.action.move_by_offset(x - self.last_x, y - self.last_y)
                self.last_x, self.last_y = x, y
                time.sleep(duration / len(points))
            
            self.action.perform()
            self.action.reset_actions()
            
        except Exception as e:
            logger.warning(f"Mouse movement to element failed: {str(e)}")

    def random_movement(self, viewport_width: int, viewport_height: int):
        """Perform random mouse movement within viewport."""
        try:
            # Calculate safe boundaries (80% of viewport)
            max_x = int(viewport_width * 0.8)
            max_y = int(viewport_height * 0.8)
            
            # Generate random target point
            target_x = random.randint(50, max_x)
            target_y = random.randint(50, max_y)
            
            # Get current position
            current_x = self.last_x if self.last_x != 0 else random.randint(0, 100)
            current_y = self.last_y if self.last_y != 0 else random.randint(0, 100)
            
            # Generate points along Bezier curve
            points = self.bezier_curve((current_x, current_y), (target_x, target_y))
            
            # Move through points
            for x, y in points:
                self.action.move_by_offset(x - self.last_x, y - self.last_y)
                self.last_x, self.last_y = x, y
                time.sleep(random.uniform(0.01, 0.03))
            
            self.action.perform()
            self.action.reset_actions()
            
        except Exception as e:
            logger.warning(f"Random mouse movement failed: {str(e)}")

class GoogleMapsScraper:
    def __init__(self, scrape_limit: int = None):
        """Initialize the scraper with optional scrape limit."""
        self.driver = None
        self.processed_urls: Set[str] = set()
        self.business_details: List[Dict] = []
        self.scrape_limit = scrape_limit
        
        try:
            self.setup_driver()
        except Exception as e:
            logger.error(f"Failed to initialize scraper: {str(e)}")
            raise

    def setup_driver(self):
        """Initialize the undetected Chrome driver with human-like settings."""
        max_retries = 3
        retry_count = 0
        
        while retry_count < max_retries:
            try:
                options = uc.ChromeOptions()
                options.add_argument('--start-maximized')
                options.add_argument('--disable-blink-features=AutomationControlled')
                options.add_argument('--no-sandbox')
                options.add_argument('--disable-dev-shm-usage')
                
                # Rotate user agent
                user_agent = random.choice(USER_AGENTS)
                options.add_argument(f'user-agent={user_agent}')
                logger.info(f"Using user agent: {user_agent}")
                
                # Add SSL certificate path
                os.environ['SSL_CERT_FILE'] = certifi.where()
                
                # Create driver with retry logic
                self.driver = uc.Chrome(options=options)
                
                # Wait for the driver to be ready
                time.sleep(2)
                
                # Set window size
                self.driver.set_window_size(1920, 1080)
                
                # Initialize mouse movement handler
                self.mouse = MouseMovement(self.driver)
                
                # Execute script after ensuring window is ready
                self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
                
                logger.info("Chrome driver initialized successfully")
                return
                
            except Exception as e:
                retry_count += 1
                logger.error(f"Attempt {retry_count} failed to setup Chrome driver: {str(e)}")
                
                # Clean up any existing driver
                try:
                    if hasattr(self, 'driver') and self.driver:
                        self.driver.quit()
                except:
                    pass
                
                if retry_count == max_retries:
                    raise Exception(f"Failed to initialize Chrome driver after {max_retries} attempts")
                
                # Wait before retrying
                time.sleep(2)

    def get_viewport_size(self) -> Tuple[int, int]:
        """Get the current viewport size."""
        viewport_width = self.driver.execute_script("return window.innerWidth;")
        viewport_height = self.driver.execute_script("return window.innerHeight;")
        return viewport_width, viewport_height

    def human_like_scroll(self, scroll_amount: int = 300):
        """Simulate human-like scrolling with random pauses."""
        current_position = self.driver.execute_script("return window.pageYOffset;")
        target_position = current_position + scroll_amount
        
        # Scroll in small increments with random pauses
        while current_position < target_position:
            scroll_step = random.randint(50, 150)
            current_position += scroll_step
            self.driver.execute_script(f"window.scrollTo(0, {current_position});")
            time.sleep(random.uniform(0.1, 0.3))

    def simulate_mouse_movement(self):
        """Simulate sophisticated mouse movements within viewport bounds."""
        try:
            viewport_width, viewport_height = self.get_viewport_size()
            self.mouse.random_movement(viewport_width, viewport_height)
        except Exception as e:
            logger.warning(f"Mouse movement simulation failed: {str(e)}")

    def extract_business_names(self) -> List[str]:
        """Extract business names from the search results."""
        try:
            # Wait for the business name elements to be present
            wait = WebDriverWait(self.driver, 10)
            elements = wait.until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.dbg0pd span.OSrXXb"))
            )
            
            # Extract text from elements
            business_names = [element.text for element in elements]
            return business_names

        except Exception as e:
            logger.error(f"Error extracting business names: {str(e)}")
            return []

    def get_next_page_url(self) -> str:
        """Extract the next page URL from the current page."""
        try:
            # Scroll to the bottom of the page to ensure pagination is loaded
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)  # Wait for any dynamic content to load

            # Try multiple selectors for the next page button
            selectors = [
                "td.d6cvqb a#pnnext",  # Primary selector based on the HTML structure
                "a#pnnext",  # Fallback to just the ID
                "td.d6cvqb a[href*='start=']",  # Look for links with start parameter
                "a[aria-label='Next page']",
                "a[aria-label='Next']"
            ]

            for selector in selectors:
                try:
                    wait = WebDriverWait(self.driver, 5)
                    next_button = wait.until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                    )
                    if next_button.is_displayed():
                        # Get the href attribute before clicking
                        next_url = next_button.get_attribute('href')
                        if next_url:
                            # Click the button to trigger the URL change
                            next_button.click()
                            time.sleep(2)  # Wait for the new URL to be set
                            return next_url
                except:
                    continue

            # If no button found, try to construct the next page URL
            current_url = self.driver.current_url
            if "start=" in current_url:
                # Extract current start value and increment it
                start_match = re.search(r'start=(\d+)', current_url)
                if start_match:
                    current_start = int(start_match.group(1))
                    next_start = current_start + 20  # Google Maps typically shows 20 results per page
                    return current_url.replace(f'start={current_start}', f'start={next_start}')
            
            logger.info("No more pages available")
            return None

        except Exception as e:
            logger.error(f"Error getting next page URL: {str(e)}")
            return None

    def click_business_card(self, element) -> bool:
        """Click on a business card and wait for the popup to load."""
        try:
            # Scroll the element into view
            self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", element)
            time.sleep(random.uniform(0.5, 1))
            
            # Move mouse to element using Bezier curve
            self.mouse.move_to_element(element)
            
            # Click the element
            element.click()
            time.sleep(random.uniform(1, 2))
            return True
        except Exception as e:
            logger.error(f"Error clicking business card: {str(e)}")
            return False

    def extract_business_details(self) -> Dict:
        """Extract detailed information from the business popup."""
        try:
            wait = WebDriverWait(self.driver, 10)
            details = {}

            # Business Name
            try:
                name_element = wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "div.Ftghae h2.qrShPb span"))
                )
                details['business_name'] = name_element.text
            except:
                details['business_name'] = "N/A"

            # Rating and Reviews
            try:
                # Wait for any rating container to be present
                wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "div.tp9Rdc, div.TLYLSe, div.PLxO5, div.NdWbqe"))
                )
                time.sleep(1)  # Give a small delay for dynamic content

                # Try multiple selectors for rating
                rating_selectors = [
                    "div.tp9Rdc span.fzTgPe",
                    "div.tp9Rdc span.fzTgPe.Aq14fc",
                    "div.TLYLSe span.yi40Hd.YrbPuc",
                    "div.PLxO5 span.fzTgPe",
                    "div.NdWbqe span.yi40Hd.YrbPuc",
                    "span[aria-label*='Rated']",
                    "span.fzTgPe",
                    "span.yi40Hd.YrbPuc"
                ]

                rating_found = False
                for selector in rating_selectors:
                    try:
                        rating_elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                        for element in rating_elements:
                            if element.is_displayed():
                                rating_text = element.text.strip()
                                if rating_text and any(c.isdigit() for c in rating_text):
                                    details['rating'] = rating_text
                                    logger.info(f"Found rating using selector '{selector}': {details['rating']}")
                                    rating_found = True
                                    break
                        if rating_found:
                            break
                    except Exception as e:
                        logger.debug(f"Selector '{selector}' failed: {str(e)}")
                        continue

                if not rating_found:
                    details['rating'] = "N/A"
                    logger.warning("Could not find rating in any format")

                # Try multiple selectors for reviews
                review_selectors = [
                    "div.tp9Rdc div.XkoHEe span.z5jxId",
                    "div.TLYLSe span.RDApEe.YrbPuc",
                    "div.PLxO5 div.XkoHEe span.z5jxId",
                    "div.NdWbqe span.RDApEe.YrbPuc",
                    "span.z5jxId",
                    "span.RDApEe.YrbPuc"
                ]

                reviews_found = False
                for selector in review_selectors:
                    try:
                        review_elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                        for element in review_elements:
                            if element.is_displayed():
                                reviews_text = element.text
                                # Remove 'reviews' and clean up
                                reviews_text = reviews_text.replace('reviews', '').strip()
                                # Remove parentheses and clean up
                                reviews_text = reviews_text.strip('()')
                                # Convert K-based numbers to numeric format
                                reviews_text = convert_review_count(reviews_text)
                                if reviews_text:
                                    details['reviews'] = reviews_text
                                    logger.info(f"Found reviews using selector '{selector}': {details['reviews']}")
                                    reviews_found = True
                                    break
                        if reviews_found:
                            break
                    except Exception as e:
                        logger.debug(f"Selector '{selector}' failed: {str(e)}")
                        continue

                if not reviews_found:
                    details['reviews'] = "N/A"
                    logger.warning("Could not find reviews in any format")

            except Exception as e:
                logger.error(f"Error extracting rating and reviews: {str(e)}")
                details['rating'] = "N/A"
                details['reviews'] = "N/A"

            # Category
            try:
                # Find all category spans
                category_spans = self.driver.find_elements(By.CSS_SELECTOR, "div.zloOqf span.YhemCb")
                if category_spans:
                    # Get the last span which should contain the actual category
                    actual_category = category_spans[-1].text
                    details['category'] = actual_category
                else:
                    details['category'] = "N/A"
            except:
                details['category'] = "N/A"

            # Website
            try:
                website_element = self.driver.find_element(By.CSS_SELECTOR, "a.n1obkb.mI8Pwc")
                details['website'] = website_element.get_attribute('href')
            except:
                details['website'] = "N/A"

            # Address
            try:
                address_element = self.driver.find_element(By.CSS_SELECTOR, "div.zloOqf span.LrzXr")
                details['address'] = address_element.text
            except:
                details['address'] = "N/A"

            # Phone
            try:
                phone_element = self.driver.find_element(By.CSS_SELECTOR, "div.zloOqf span.LrzXr.zdqRlf a")
                phone_number = phone_element.text
                details['phone'] = format_phone_number(phone_number)
            except:
                details['phone'] = "N/A"

            return details

        except Exception as e:
            logger.error(f"Error extracting business details: {str(e)}")
            return {}

    def scrape_businesses(self, start_url: str) -> List[Dict]:
        """Main method to scrape business details from all pages of Google Maps search results."""
        all_business_details = []
        current_url = start_url
        
        try:
            while current_url and current_url not in self.processed_urls:
                logger.info(f"Processing URL: {current_url}")
                self.processed_urls.add(current_url)
                
                # Navigate to the page
                self.driver.get(current_url)
                
                # Wait for initial page load
                time.sleep(random.uniform(2, 4))
                
                # Simulate initial mouse movement
                self.simulate_mouse_movement()
                
                # Scroll through the page multiple times
                for _ in range(3):
                    # Add more randomization to scrolling
                    scroll_amount = random.randint(200, 400)
                    self.human_like_scroll(scroll_amount)
                    time.sleep(random.uniform(1, 2))
                    self.simulate_mouse_movement()
                
                # Find all business cards
                wait = WebDriverWait(self.driver, 10)
                business_cards = wait.until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.dbg0pd"))
                )
                
                # Process each business card
                for card in business_cards:
                    # Check if we've reached the scrape limit
                    if self.scrape_limit and len(all_business_details) >= self.scrape_limit:
                        logger.info(f"Reached scrape limit of {self.scrape_limit} businesses")
                        return all_business_details

                    if self.click_business_card(card):
                        # Extract details from the popup
                        details = self.extract_business_details()
                        if details:
                            all_business_details.append(details)
                            logger.info(f"Extracted details for: {details.get('business_name', 'Unknown')}")
                        
                        # Add a random delay between cards
                        time.sleep(random.uniform(1, 2))
                
                logger.info(f"Extracted {len(business_cards)} business details from current page")
                
                # Check if we've reached the scrape limit before getting next page
                if self.scrape_limit and len(all_business_details) >= self.scrape_limit:
                    logger.info(f"Reached scrape limit of {self.scrape_limit} businesses")
                    return all_business_details
                
                # Get next page URL
                current_url = self.get_next_page_url()
                
                # Add a random delay between pages
                time.sleep(random.uniform(3, 7))
            
            logger.info(f"Successfully extracted {len(all_business_details)} total business details")
            return all_business_details

        except Exception as e:
            logger.error(f"Error during scraping: {str(e)}")
            return all_business_details
        
        finally:
            if self.driver:
                self.driver.quit()

def create_search_url(search_query: str) -> str:
    """Create a Google Maps search URL from the search query."""
    base_url = "https://www.google.com/search"
    params = {
        "tbm": "lcl",  # Local search
        "q": search_query,
        "rflfq": "1",
        "num": "10",
        "sa": "X"
    }
    return f"{base_url}?{urllib.parse.urlencode(params)}"

def create_excel_file(business_details: List[Dict], search_query: str) -> str:
    """Create an Excel file with the business details."""
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        
        # Set headers in the new sequence
        headers = [
            'Business Name',
            'Rating',
            'Reviews',
            'Category',
            'Phone',
            'Website',
            'Address'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(name='Tahoma', size=11)
            cell.alignment = Alignment(vertical='center')
        
        # Add data in the new sequence
        for row, details in enumerate(business_details, 2):
            ws.cell(row=row, column=1, value=details.get('business_name', 'N/A'))
            ws.cell(row=row, column=2, value=details.get('rating', 'N/A'))
            ws.cell(row=row, column=3, value=details.get('reviews', 'N/A'))
            ws.cell(row=row, column=4, value=details.get('category', 'N/A'))
            ws.cell(row=row, column=5, value=details.get('phone', 'N/A'))
            ws.cell(row=row, column=6, value=details.get('website', 'N/A'))
            ws.cell(row=row, column=7, value=details.get('address', 'N/A'))
        
        # Set row height and formatting for all rows
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 21
            for cell in row:
                cell.font = Font(name='Tahoma', size=11)
                cell.alignment = Alignment(vertical='center')
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].auto_size = True
        
        # Create filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"business_details_{search_query.replace(' ', '_')}_{timestamp}.xlsx"
        
        # Save the workbook
        wb.save(filename)
        print(f"\nExcel file created successfully: {filename}")
        return filename
        
    except Exception as e:
        print(f"Error creating Excel file: {str(e)}")
        return None

def main():
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Scrape business details from Google Maps')
    parser.add_argument('search_query', type=str, help='Search query for Google Maps')
    parser.add_argument('--limit', type=int, help='Limit the number of businesses to scrape')
    args = parser.parse_args()

    # Create search URL
    start_url = create_search_url(args.search_query)
    
    # Initialize scraper with limit
    scraper = GoogleMapsScraper(scrape_limit=args.limit)
    business_details = scraper.scrape_businesses(start_url)
    
    # Save results to a JSON file
    output_file = f"business_details_{args.search_query.replace(' ', '_')}.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(business_details, f, indent=2, ensure_ascii=False)
    
    # Create Excel file
    excel_file = create_excel_file(business_details, args.search_query)
    if excel_file:
        print(f"Data has been saved to: {excel_file}")
    else:
        print("Failed to create Excel file. Data has been saved to JSON file only.")

if __name__ == "__main__":
    main() 